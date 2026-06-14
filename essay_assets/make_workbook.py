#!/usr/bin/env python3
"""Build a native Excel workbook (.xlsx) with the essay's data tables and Excel charts."""
import os
from openpyxl import Workbook
from openpyxl.chart import LineChart, BarChart, ScatterChart, Reference, Series
from openpyxl.chart.trendline import Trendline
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = os.path.dirname(os.path.abspath(__file__))

wb = Workbook()

HEAD = Font(bold=True, color="FFFFFF", size=11)
HFILL = PatternFill("solid", fgColor="1F3B63")
TITLE = Font(bold=True, size=14, color="1F3B63")
NOTE = Font(italic=True, size=9, color="7F8C8D")
thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center")


def style_header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEAD
        cell.fill = HFILL
        cell.alignment = CENTER
        cell.border = BORDER


# ---------------- Sheet 1: Minimum wage & non-regular share ----------------
ws = wb.active
ws.title = "MinWage_NonRegular"
ws["A1"] = "Japan: Minimum Wage and Non-Regular Employment Share, 2005-2025"
ws["A1"].font = TITLE
ws["A2"] = "Sources: MHLW Regional Minimum Wage Survey; Statistics Bureau, Labour Force Survey (Detailed Tabulation)."
ws["A2"].font = NOTE

hdr = ["Year", "Minimum wage (¥/hr)", "YoY change (%)", "Non-regular share (%)"]
ws.append([])
ws.append(hdr)
style_header(ws, 4, len(hdr))

years = list(range(2005, 2026))
minwage = [668, 673, 687, 703, 713, 730, 737, 749, 764, 780, 798,
           823, 848, 874, 901, 902, 930, 961, 1004, 1055, 1121]
nonreg = {y: v for y, v in zip(range(2005, 2024),
          [32.6, 33.0, 33.5, 34.1, 33.7, 34.4, 35.1, 35.2, 36.7, 37.4,
           37.5, 37.5, 37.3, 38.2, 38.3, 37.2, 36.7, 36.9, 37.1])}
for i, y in enumerate(years):
    yoy = "" if i == 0 else round((minwage[i] / minwage[i - 1] - 1) * 100, 1)
    ws.append([y, minwage[i], yoy, nonreg.get(y, "")])
for r in range(5, 5 + len(years)):
    for c in range(1, 5):
        ws.cell(row=r, column=c).border = BORDER
        ws.cell(row=r, column=c).alignment = CENTER
for col, w in zip("ABCD", [8, 20, 16, 22]):
    ws.column_dimensions[col].width = w

first, last = 5, 4 + len(years)
# Line chart: minimum wage
lc = LineChart()
lc.title = "Minimum Wage (¥/hour), 2005-2025"
lc.y_axis.title = "¥/hour"
lc.x_axis.title = "Year"
lc.height, lc.width = 9, 16
data = Reference(ws, min_col=2, min_row=4, max_row=last)
cats = Reference(ws, min_col=1, min_row=first, max_row=last)
lc.add_data(data, titles_from_data=True)
lc.set_categories(cats)
lc.series[0].graphicalProperties.line.solidFill = "1F3B63"
lc.series[0].graphicalProperties.line.width = 28000
ws.add_chart(lc, "F4")

# Line chart: non-regular share
lc2 = LineChart()
lc2.title = "Non-Regular Share (%), 2005-2023"
lc2.y_axis.title = "%"
lc2.x_axis.title = "Year"
lc2.height, lc2.width = 9, 16
data2 = Reference(ws, min_col=4, min_row=4, max_row=4 + 19)  # 2005-2023
cats2 = Reference(ws, min_col=1, min_row=first, max_row=4 + 19)
lc2.add_data(data2, titles_from_data=True)
lc2.set_categories(cats2)
lc2.series[0].graphicalProperties.line.solidFill = "2A9D8F"
lc2.series[0].graphicalProperties.line.width = 28000
ws.add_chart(lc2, "F23")

# ---------------- Sheet 2: Prefectural minimum wages 2024 ----------------
ws2 = wb.create_sheet("Prefectural_2024")
ws2["A1"] = "Hourly Minimum Wage by Prefecture, 2024 (selected)"
ws2["A1"].font = TITLE
ws2["A2"] = "Source: MHLW, October 2024 revision. National weighted average = ¥1,055."
ws2["A2"].font = NOTE
ws2.append([])
ws2.append(["Prefecture", "Minimum wage (¥/hr)"])
style_header(ws2, 4, 2)
pref = [("Tokyo", 1163), ("Kanagawa", 1162), ("Osaka", 1114), ("Saitama", 1078),
        ("Aichi", 1077), ("Chiba", 1076), ("Kyoto", 1058), ("Hyogo", 1052),
        ("Hiroshima", 1020), ("Hokkaido", 1010), ("Fukuoka", 992), ("Miyagi", 973),
        ("Aomori", 953), ("Iwate", 952), ("Okinawa", 952), ("Akita", 951)]
for name, v in pref:
    ws2.append([name, v])
for r in range(5, 5 + len(pref)):
    for c in range(1, 3):
        ws2.cell(row=r, column=c).border = BORDER
ws2.column_dimensions["A"].width = 14
ws2.column_dimensions["B"].width = 20
bc = BarChart()
bc.type = "bar"
bc.title = "Minimum Wage by Prefecture, 2024"
bc.x_axis.title = "Prefecture"
bc.y_axis.title = "¥/hour"
bc.height, bc.width = 11, 18
d = Reference(ws2, min_col=2, min_row=4, max_row=4 + len(pref))
c = Reference(ws2, min_col=1, min_row=5, max_row=4 + len(pref))
bc.add_data(d, titles_from_data=True)
bc.set_categories(c)
bc.series[0].graphicalProperties.solidFill = "2E6DA4"
bc.legend = None
ws2.add_chart(bc, "D4")

# ---------------- Sheet 3: Elasticities ----------------
ws3 = wb.create_sheet("Elasticities")
ws3["A1"] = "Estimated Employment Elasticities of the Minimum Wage"
ws3["A1"].font = TITLE
ws3["A2"] = "Elasticity = % change in employment per 1% increase in the minimum wage."
ws3["A2"].font = NOTE
ws3.append([])
ws3.append(["Study", "Group / setting", "Elasticity"])
style_header(ws3, 4, 3)
el = [
    ("Kawaguchi & Mori (2021)", "Young less-educated men, Japan", -1.2),
    ("Abe (2011)", "Teenagers, Japan", -0.4),
    ("This study (expected)", "Japan aggregate", -0.3),
    ("Neumark & Wascher (2000)", "US teenagers", -0.15),
    ("Dube, Lester & Reich (2010)", "US, contiguous border counties", -0.02),
]
for s, g, v in el:
    ws3.append([s, g, v])
for r in range(5, 5 + len(el)):
    for c in range(1, 4):
        ws3.cell(row=r, column=c).border = BORDER
ws3.column_dimensions["A"].width = 28
ws3.column_dimensions["B"].width = 32
ws3.column_dimensions["C"].width = 12
bc2 = BarChart()
bc2.type = "bar"
bc2.title = "Employment Elasticity by Study"
bc2.y_axis.title = "Elasticity"
bc2.height, bc2.width = 9, 18
d2 = Reference(ws3, min_col=3, min_row=4, max_row=4 + len(el))
c2 = Reference(ws3, min_col=1, min_row=5, max_row=4 + len(el))
bc2.add_data(d2, titles_from_data=True)
bc2.set_categories(c2)
bc2.series[0].graphicalProperties.solidFill = "C0392B"
bc2.legend = None
ws3.add_chart(bc2, "E4")

# ---------------- Sheet 4: Prefectural correlation (all 47) ----------------
ws4 = wb.create_sheet("Prefecture_Correlation")
ws4["A1"] = "Minimum Wage vs. Non-Regular Employment, All 47 Prefectures"
ws4["A1"].font = TITLE
ws4["A2"] = ("Sources: MHLW FY2025 minimum-wage table (2024 & 2025); 2022 Employment Status Survey, "
             "Table 2-3 (non-regular share of all workers).")
ws4["A2"].font = NOTE
ws4.append([])
ws4.append(["Prefecture", "Min wage 2024 (¥)", "Min wage 2025 (¥)", "Non-regular share 2022 (%)"])
style_header(ws4, 4, 4)
pref_panel = [
    ("Hokkaido",1010,1075,34.3),("Aomori",953,1029,29.1),("Iwate",952,1031,29.7),("Miyagi",973,1038,30.3),
    ("Akita",951,1031,28.8),("Yamagata",955,1032,26.8),("Fukushima",955,1033,27.8),("Ibaraki",1005,1074,31.8),
    ("Tochigi",1004,1068,31.0),("Gunma",985,1063,32.2),("Saitama",1078,1141,33.5),("Chiba",1076,1140,32.4),
    ("Tokyo",1163,1226,28.0),("Kanagawa",1162,1225,32.2),("Niigata",985,1050,29.3),("Toyama",998,1062,27.7),
    ("Ishikawa",984,1054,29.2),("Fukui",984,1053,28.1),("Yamanashi",988,1052,31.4),("Nagano",998,1061,30.0),
    ("Gifu",1001,1065,32.7),("Shizuoka",1034,1097,32.4),("Aichi",1077,1140,32.1),("Mie",1023,1087,33.4),
    ("Shiga",1017,1080,35.0),("Kyoto",1058,1122,34.2),("Osaka",1114,1177,34.1),("Hyogo",1052,1116,33.9),
    ("Nara",986,1051,34.5),("Wakayama",980,1045,30.4),("Tottori",957,1030,29.2),("Shimane",962,1033,30.4),
    ("Okayama",982,1047,30.0),("Hiroshima",1020,1085,31.2),("Yamaguchi",979,1043,30.8),("Tokushima",980,1046,26.6),
    ("Kagawa",970,1036,28.8),("Ehime",956,1033,29.0),("Kochi",952,1023,28.3),("Fukuoka",992,1057,34.2),
    ("Saga",956,1030,30.5),("Nagasaki",953,1031,32.2),("Kumamoto",952,1034,29.9),("Oita",954,1035,29.6),
    ("Miyazaki",952,1023,30.7),("Kagoshima",953,1026,31.6),("Okinawa",952,1023,33.4),
]
for row in pref_panel:
    ws4.append(list(row))
first4, last4 = 5, 4 + len(pref_panel)
for r in range(first4, last4 + 1):
    for c in range(1, 5):
        ws4.cell(row=r, column=c).border = BORDER
for col, w in zip("ABCD", [13, 18, 18, 24]):
    ws4.column_dimensions[col].width = w
# Pearson correlation in-sheet
ws4["F4"] = "Pearson r (2024 wage vs non-reg):"
ws4["I4"] = f"=CORREL(B{first4}:B{last4},D{first4}:D{last4})"
ws4["F5"] = "R-squared:"
ws4["I5"] = f"=I4^2"
for cell in ("F4", "F5"):
    ws4[cell].font = Font(bold=True)
# Scatter chart: x = min wage 2024, y = non-reg share
sc = ScatterChart()
sc.title = "Min Wage (2024) vs Non-Regular Share (2022), 47 Prefectures"
sc.x_axis.title = "Minimum wage 2024 (¥/hour)"
sc.y_axis.title = "Non-regular share 2022 (%)"
sc.x_axis.scaling.min = 940; sc.x_axis.scaling.max = 1180
sc.y_axis.scaling.min = 26; sc.y_axis.scaling.max = 36
sc.height, sc.width = 11, 18
xref = Reference(ws4, min_col=2, min_row=first4, max_row=last4)
yref = Reference(ws4, min_col=4, min_row=first4, max_row=last4)
series = Series(yref, xref, title="Prefectures")
series.marker.symbol = "circle"; series.marker.size = 6
series.graphicalProperties.line.noFill = True
series.trendline = Trendline(trendlineType="linear", dispRSqr=True, dispEq=True)
sc.series.append(series)
ws4.add_chart(sc, "F8")

path = os.path.join(OUT, "Japan_MinimumWage_Data.xlsx")
wb.save(path)
print("Workbook saved:", path)
